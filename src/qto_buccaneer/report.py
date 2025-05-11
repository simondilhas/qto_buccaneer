import pandas as pd
import os
from pathlib import Path
from typing import Optional, Dict, Any, Union
from jinja2 import Environment, FileSystemLoader
import subprocess
from dataclasses import dataclass
import openpyxl
from openpyxl.utils import get_column_letter
from weasyprint import HTML
from datetime import datetime
import pdfkit
import yaml
from qto_buccaneer._utils.report.excel_styling import ExcelLayoutConfig
from qto_buccaneer._utils.report.pdf_styling import ReportStyleConfig
from qto_buccaneer._utils.report.project_comparision_table import _export_project_comparison_excel, _create_project_comparison_df
from qto_buccaneer._utils.report.generate_metrics_report import _convert_html_to_pdf, _build_metrics_table


def room_program_comparison(
    target_excel_path: str,
    ifc_loader,
    room_name_column: str = "LongName",
    target_count_column: Optional[str] = None,
    target_area_column: str = "Target Area/Room",
    ifc_room_name_attribute: str = "LongName",
    ifc_area_attribute: str = "Qto_SpaceBaseQuantities.NetFloorArea",
    output_path: Optional[str] = None,
    layout_config: Optional[ExcelLayoutConfig] = None
) -> pd.DataFrame:
    """
    Create a comparison between target room program and actual IFC spaces.
    
    The function expects a target room program in Excel format with the following structure:
    - One row per room type (e.g., "Office", "Meeting Room", "Bathroom")
    - Columns for room type name and target area per room
    - Optionally a column for target count, or it will be calculated from room names
    
    Args:
        target_excel_path: Path to Excel file containing target room program.
            Expected format: One row per room type with columns for name and area.
        ifc_loader: Instance of IfcLoader with loaded IFC model
        room_name_column: Column name in target Excel for room types (default: "LongName")
        target_count_column: Optional column name for target room count. If None, will be calculated.
        target_area_column: Column name for target area per room (default: "Target Area/Room")
        ifc_room_name_attribute: Attribute name in IFC for room names (default: "LongName")
        ifc_area_attribute: Attribute name in IFC for area values (default: "Qto_SpaceBaseQuantities.NetFloorArea")
        output_path: Optional path to save Excel report
        layout_config: Optional ExcelLayoutConfig for custom formatting
        
    Returns:
        pd.DataFrame: Comparison table with the following metrics for each room type:
            - Target Count: Number of rooms planned (calculated if not provided)
            - Target Area/Room: Planned area per room
            - Target Total Area: Total planned area (count × area/room)
            - Actual Count: Number of rooms in IFC
            - Actual Total Area: Total area in IFC
            - Average Area/Room: Actual average area per room
            - Count Difference: Actual - Target count
            - Area Difference: Actual - Target total area
            - Percentage differences for both count and area
    """
    # Load target room program
    try:
        target_df = pd.read_excel(target_excel_path)
        print(f"Loaded target Excel file. Columns found: {target_df.columns.tolist()}")  # Debug print
        
        # Verify required columns exist
        missing_columns = []
        for col, col_name in [
            (room_name_column, "room name"),
            (target_area_column, "target area")
        ]:
            if col not in target_df.columns:
                missing_columns.append(f"{col_name} ({col})")
        
        if missing_columns:
            raise ValueError(f"Missing required columns in target Excel: {', '.join(missing_columns)}")
            
        # If no target_count_column provided, calculate counts from room names
        if target_count_column is None:
            print("No target count column provided - calculating counts from room names")
            # Group by room name and count occurrences
            count_df = target_df.groupby(room_name_column).size().reset_index(name='Target Count')
            # Merge with original DataFrame
            target_df = pd.merge(target_df, count_df, on=room_name_column)
            target_count_column = 'Target Count'
        elif target_count_column not in target_df.columns:
            raise ValueError(f"Target count column '{target_count_column}' not found in Excel")
            
    except Exception as e:
        print(f"Error loading target Excel file: {str(e)}")
        return pd.DataFrame()
        
    # Get actual spaces from IFC
    try:
        spaces_df = ifc_loader.get_space_information()
        print(f"Loaded spaces from IFC. Found {len(spaces_df)} spaces.")  # Debug print
        
        if spaces_df.empty:
            raise ValueError("No spaces found in IFC model")
            
        # Verify required IFC data columns
        if ifc_room_name_attribute not in spaces_df.columns:
            raise ValueError(f"IFC spaces missing '{ifc_room_name_attribute}' attribute")
        if ifc_area_attribute not in spaces_df.columns:
            raise ValueError(f"IFC spaces missing '{ifc_area_attribute}' quantity")
            
    except Exception as e:
        print(f"Error processing IFC spaces: {str(e)}")
        return pd.DataFrame()
    
    # Initialize result DataFrame
    result = pd.DataFrame()
    
    try:
        # Process each room type from target program
        data = []
        for _, row in target_df.iterrows():
            room_name = row[room_name_column]
            target_count = float(row[target_count_column])  # Convert to float for safety
            target_area = float(row[target_area_column])
            
            print(f"Processing room type: {room_name}")  # Debug print
            
            # Get actual spaces matching this room name
            actual_spaces = spaces_df[spaces_df[ifc_room_name_attribute] == room_name]
            actual_count = len(actual_spaces)
            
            # Sum up actual areas
            actual_total_area = actual_spaces[ifc_area_attribute].sum()
            
            print(f"Found {actual_count} spaces with total area {actual_total_area}")  # Debug print
            
            # Calculate metrics
            target_total_area = target_count * target_area
            avg_area_per_room = actual_total_area / actual_count if actual_count > 0 else 0
            
            count_diff = actual_count - target_count
            count_diff_pct = (count_diff / target_count * 100) if target_count > 0 else 0
            
            area_diff = actual_total_area - target_total_area
            area_diff_pct = (area_diff / target_total_area * 100) if target_total_area > 0 else 0
            
            data.append({
                'Room Type': room_name,
                'Target Count': target_count,
                'Target sqm/room': target_area,
                'Target Total sqm': target_total_area,
                'Actual Count': actual_count,
                'Actual Total sqm': actual_total_area,
                'Avg sqm/room': avg_area_per_room,
                'Count Diff': count_diff,
                '% Count Diff': count_diff_pct,
                'Area Diff': area_diff,
                '% Area Diff': area_diff_pct
            })
            
        result = pd.DataFrame(data)
        
        if result.empty:
            print("No data was processed - empty result DataFrame")
            return pd.DataFrame()
            
        # Add totals row
        totals = {
            'Room Type': 'TOTAL',
            'Target Count': result['Target Count'].sum(),
            'Target Total sqm': result['Target Total sqm'].sum(),
            'Actual Count': result['Actual Count'].sum(),
            'Actual Total sqm': result['Actual Total sqm'].sum(),
            'Count Diff': result['Count Diff'].sum(),
            'Area Diff': result['Area Diff'].sum()
        }
        
        # Calculate weighted averages for percentages
        total_target_count = result['Target Count'].sum()
        total_target_area = result['Target Total sqm'].sum()
        
        if total_target_count > 0:
            totals['% Count Diff'] = (totals['Count Diff'] / total_target_count * 100)
        if total_target_area > 0:
            totals['% Area Diff'] = (totals['Area Diff'] / total_target_area * 100)
            
        # Calculate overall average sqm/room
        if totals['Actual Count'] > 0:
            totals['Avg sqm/room'] = totals['Actual Total sqm'] / totals['Actual Count']
        else:
            totals['Avg sqm/room'] = 0
            
        totals['Target sqm/room'] = totals['Target Total sqm'] / totals['Target Count'] if totals['Target Count'] > 0 else 0
        
        # Append totals row
        result = pd.concat([result, pd.DataFrame([totals])], ignore_index=True)
        
        # Export to Excel if output path is provided
        if output_path:
            result = _export_room_program_comparison(
                df=result,
                output_path=output_path,
                layout_config=layout_config
            )
        
        return result
        
    except Exception as e:
        print(f"Error creating comparison: {str(e)}")
        import traceback
        traceback.print_exc()  # Print full stack trace
        return pd.DataFrame()



# TODO: adapt to new report style with class ReportResults
def generate_metrics_report(
    metrics_df: pd.DataFrame,
    building_name: str,
    plots_dir: Union[str, Path],
    building_adresse: Optional[str] = None,
    building_description: Optional[str] = None,
    output_dir: Union[str, Path] = 'reports',
    template_path: Union[str, Path] = 'configs/abstractBIM_report_template.html',
    style_config: Optional[ReportStyleConfig] = None,
    report_config_path: Union[str, Path] = "abstractBIM_report_config.yaml"
) -> str:
    """
    Generate a metrics report from the provided metrics DataFrame.
    
    Args:
        metrics_df (pd.DataFrame): DataFrame containing the metrics data
        building_name (str): Name of the project, used for the report title
        plots_dir (Union[str, Path]): Directory containing report plots
        building_adresse (Optional[str]): Address of the building
        building_description (Optional[str]): Description of the building
        output_dir (Union[str, Path]): Directory where the final PDF report will be saved
        template_path (Union[str, Path]): Path to the report template HTML file
        style_config (Optional[ReportStyleConfig]): Configuration for report styling
        report_config_path (Union[str, Path]): Path to the report configuration YAML file
        
    Returns:
        str: Path to the generated PDF report
    """
    # Convert all paths to Path objects
    plots_dir = Path(plots_dir)
    output_dir = Path(output_dir)
    template_path = Path(template_path)
    report_config_path = Path(report_config_path)
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Construct output file path
    output_file = f"{building_name}_report.pdf"
    output_path = os.path.join(output_dir, output_file)
    
    print(f"Looking for template at: {template_path}")  # Debug print
    
    # Set default image placeholders and formats
    image_placeholders = [
        'titel_picture'
        ]
    image_formats = ['.png', '.jpg', '.jpeg']
    
    # Collect available images
    images = {}
    for key in image_placeholders:
        found = False
        for ext in image_formats:
            img_path = os.path.join(plots_dir, f"{key}{ext}")
            print(f"Looking for image at: {img_path}")  # Debug print
            if os.path.isfile(img_path):
                print(f"Found image: {img_path}")  # Debug print
                # Convert absolute path to relative path
                # Calculate relative path from the HTML file's directory to the image
                rel_path = os.path.relpath(img_path, os.path.dirname(output_path))
                # Convert Windows backslashes to forward slashes for web compatibility
                rel_path = rel_path.replace('\\', '/')
                images[key] = rel_path
                found = True
                break
        if not found:
            print(f"Image not found for key: {key}")  # Debug print
            images[key] = None
    
    # Verify template file exists
    if not os.path.isfile(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    # Load report configuration if provided
    include_metrics = None
    if report_config_path:
        try:
            with open(report_config_path, 'r') as f:
                report_config = yaml.safe_load(f)
                include_metrics = report_config.get('include_metrics', [])
                print(f"Loaded include_metrics from config: {include_metrics}")  # Debug print
        except Exception as e:
            print(f"Warning: Could not load report config: {e}")
    
    # Create metrics table using the build_metrics_table function
    metrics_table = _build_metrics_table(metrics_df, config_path=report_config_path)
    
    # Render HTML
    template_dir = os.path.dirname(template_path)
    template_name = os.path.basename(template_path)
    env = Environment(loader=FileSystemLoader(template_dir))
    template = env.get_template(template_name)
    
    html_out = template.render(
        project_name=building_name,
        file_name=building_name,
        address=building_adresse or "",
        description=building_description or "",
        date_time=datetime.now().strftime('%Y-%m-%d %H:%M'),
        images=images,
        logo_path=style_config.logo_path if style_config else None,
        metrics_table=metrics_table  # Pass the metrics table to the template
    )
    
    # Save HTML
    html_path = output_path.replace('.pdf', '.html')
    with open(html_path, 'w') as f:
        f.write(html_out)
    
    # Convert to PDF with styling
    try:
        _convert_html_to_pdf(html_out, output_path, style_config)
    except Exception as e:
        print(f"Warning: Could not convert to PDF: {e}")
        print(f"HTML report saved at: {html_path}")
        return html_path
    
    return output_path



def generate_project_comparison(
    df: pd.DataFrame,
    output_path: Optional[str] = None,
    include_metrics: Optional[list[str]] = None,
    layout_config: Optional[ExcelLayoutConfig] = None
) -> pd.DataFrame:
    """
    Create and export a project comparison DataFrame.
    
    Args:
        df (pd.DataFrame): Input DataFrame containing metrics data
        output_path (Optional[str]): Optional path to save Excel report
        include_metrics (Optional[list[str]]): List of metric names to include in the comparison
        layout_config (Optional[ExcelLayoutConfig]): Configuration for Excel layout.
            If None, default settings will be used.
            
    Returns:
        pd.DataFrame: Comparison table with projects as rows and metrics as columns
    """
    try:
        # Create the comparison DataFrame
        result = _create_project_comparison_df(df, include_metrics)
        
        if result.empty:
            print("No data was processed - empty result DataFrame")
            return pd.DataFrame()
            
        # Export to Excel if output path is provided
        if output_path:
            result = _export_project_comparison_excel(
                df=result,
                output_path=output_path,
                include_metrics=include_metrics,
                layout_config=layout_config
            )
        
        return result
        
    except Exception as e:
        print(f"Error creating comparison: {str(e)}")
        import traceback
        traceback.print_exc()  # Print full stack trace
        return pd.DataFrame()

def generate_room_program_from_excel(
    input_excel_path: str,
    room_name_column: str = "LongName",
    area_column: str = "NetFloorArea",
    count_column: Optional[str] = None,
    output_path: Optional[str] = None,
    layout_config: Optional[ExcelLayoutConfig] = None
) -> pd.DataFrame:
    """
    Convert an Excel file with individual rooms into an aggregated room program format.
    
    This function takes an Excel file containing individual rooms and aggregates them by room type,
    calculating the count and average area for each type. The output format matches
    what is expected by room_program_comparison().
    
    Args:
        input_excel_path: Path to Excel file containing individual rooms.
            Each row should represent one room with at least:
            - A room type/name column (default: "LongName")
            - An area column (default: "NetFloorArea")
            - Optionally a count column if rooms are already grouped
        room_name_column: Column name in input Excel for the room type/name
        area_column: Column name in input Excel for the area value
        count_column: Optional column name for room count. If provided, uses this instead of counting rows.
        output_path: Optional path to save the aggregated room program as Excel
        layout_config: Optional ExcelLayoutConfig for custom formatting
        
    Returns:
        pd.DataFrame: Aggregated room program with columns:
            - Room Type (from room_name_column)
            - Target Count (number of rooms of each type)
            - Target Area/Room (average area per room type)
            
    Example:
        Input Excel format:
        | LongName | Soll m2 | Soll Anzahl | Other columns... |
        |----------|---------|-------------|------------------|
        | Office   | 20.0    | 2           | ...              |
        | Meeting  | 30.0    | 1           | ...              |
        
        Output DataFrame:
        | Room Type | Target Count | Target Area/Room |
        |-----------|--------------|------------------|
        | Office    | 2           | 20.0            |
        | Meeting   | 1           | 30.0            |
    """
    try:
        # Load input Excel file
        df = pd.read_excel(input_excel_path)
        
        if df.empty:
            raise ValueError("Input Excel file is empty")
            
        # Verify required columns exist
        missing_columns = []
        for col, col_name in [
            (room_name_column, "room name"),
            (area_column, "area")
        ]:
            if col not in df.columns:
                missing_columns.append(f"{col_name} ({col})")
                
        if missing_columns:
            raise ValueError(f"Missing required columns in input Excel: {', '.join(missing_columns)}")
            
        # If count_column is provided, use it directly
        if count_column:
            if count_column not in df.columns:
                raise ValueError(f"Count column '{count_column}' not found in input Excel")
                
            # Group by room type and use provided count
            result = df.groupby(room_name_column).agg({
                count_column: 'sum',
                area_column: 'mean'
            }).reset_index()
            
            # Rename columns to match expected output format
            result.columns = [room_name_column, 'Target Count', 'Target Area/Room']
            
        else:
            # Group by room type and calculate metrics
            result = df.groupby(room_name_column).agg({
                area_column: ['count', 'mean']
            }).reset_index()
            
            # Flatten multi-index columns
            result.columns = [room_name_column, 'Target Count', 'Target Area/Room']
        
        # Round area to 2 decimal places
        result['Target Area/Room'] = result['Target Area/Room'].round(2)
        
        # Export to Excel if path provided
        if output_path:
            # Create output directory if it doesn't exist
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
                
            # Use default config if none provided
            config = layout_config or ExcelLayoutConfig(
                horizontal_lines=True,
                vertical_lines=True,
                bold_headers=True,
                auto_column_width=True
            )
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result.to_excel(writer, index=False, sheet_name='Room Program')
                worksheet = writer.sheets['Room Program']
                
                # Apply styling based on config
                if config.bold_headers:
                    for cell in worksheet[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        if config.header_color:
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color=config.header_color,
                                end_color=config.header_color,
                                fill_type='solid'
                            )
                
                # Set number format
                number_format = config.number_format
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = number_format
                
                # Auto-adjust column widths
                if config.auto_column_width:
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        return result
        
    except Exception as e:
        print(f"Error creating room program: {str(e)}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()

def generate_roomprogram_gropued_by_name(
    input_excel_path: str,
    input_room_name_column: str = "LongName",
    input_area_column: str = "Soll m2",
    output_room_name_column: str = "Room Type",
    output_area_column: str = "Area",
    output_count_column: str = "Soll Anzahl",
    output_path: Optional[str] = None,
    layout_config: Optional[ExcelLayoutConfig] = None
) -> pd.DataFrame:
    """
    Group rooms by name and calculate aggregated metrics.
    
    This function takes an Excel file containing individual rooms and:
    1. Groups them by room name
    2. Counts the number of rooms in each group (minimum 1)
    3. Sums the area values for each group (NaN for rooms without area)
    
    Args:
        input_excel_path: Path to Excel file containing individual rooms
        input_room_name_column: Column name for room types/names in input file
        input_area_column: Column name for area values in input file
        output_room_name_column: Column name for room types/names in output
        output_area_column: Column name for area values in output
        output_count_column: Column name for count values in output
        output_path: Optional path to save the results as Excel
        layout_config: Optional ExcelLayoutConfig for custom formatting
        
    Returns:
        pd.DataFrame: Aggregated room program with columns:
            - {output_room_name_column} (from input_room_name_column)
            - {output_area_column} (sum of areas for each room type, NaN for rooms without area)
            - {output_count_column} (number of rooms of each type, minimum 1)
    """
    try:
        # Load input Excel file
        df = pd.read_excel(input_excel_path)
        
        if df.empty:
            raise ValueError("Input Excel file is empty")
            
        # Verify required columns exist
        missing_columns = []
        for col, col_name in [
            (input_room_name_column, "room name"),
            (input_area_column, "area")
        ]:
            if col not in df.columns:
                missing_columns.append(f"{col_name} ({col})")
                
        if missing_columns:
            raise ValueError(f"Missing required columns in input Excel: {', '.join(missing_columns)}")
            
        # Group by room name and calculate metrics
        result = df.groupby(input_room_name_column).agg({
            input_area_column: lambda x: x.sum() if not x.isna().all() else pd.NA  # Sum areas, return NA if all are NA
        }).reset_index()
        
        # Add count column (count all rooms, minimum 1)
        result[output_count_column] = df.groupby(input_room_name_column).size().clip(lower=1).values
        
        # Rename columns for clarity
        result.columns = [output_room_name_column, output_area_column, output_count_column]
        
        # Export to Excel if path provided
        if output_path:
            # Create output directory if it doesn't exist
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
                
            # Use default config if none provided
            config = layout_config or ExcelLayoutConfig(
                horizontal_lines=True,
                vertical_lines=True,
                bold_headers=True,
                auto_column_width=True
            )
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result.to_excel(writer, index=False, sheet_name='Room Program')
                worksheet = writer.sheets['Room Program']
                
                # Apply styling based on config
                if config.bold_headers:
                    for cell in worksheet[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        if config.header_color:
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color=config.header_color,
                                end_color=config.header_color,
                                fill_type='solid'
                            )
                
                # Set number format
                number_format = config.number_format
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = number_format
                
                # Auto-adjust column widths
                if config.auto_column_width:
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        return result
        
    except Exception as e:
        print(f"Error grouping rooms: {str(e)}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()


