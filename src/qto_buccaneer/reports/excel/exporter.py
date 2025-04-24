import pandas as pd
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side

from ..config import ExcelLayoutConfig

def export_to_excel(
    df: pd.DataFrame, 
    path: str, 
    config: Optional[ExcelLayoutConfig] = None,
    suffix: Optional[str] = None
) -> None:
    """Export a DataFrame to a new Excel file with optional styling.
    
    Args:
        df (pd.DataFrame): DataFrame to export
        path (str): Path where the Excel file should be saved
        config (Optional[ExcelLayoutConfig]): Configuration for Excel styling. 
            If None, uses default ExcelLayoutConfig settings.
        suffix (Optional[str]): Optional suffix to append to the filename before the .xlsx extension
    """
    if df.empty:
        return
        
    # Use default config if none provided
    if config is None:
        config = ExcelLayoutConfig()
    
    # Add suffix to path if provided
    if suffix:
        path = str(Path(path).with_suffix('')) + suffix + '.xlsx'
    
    # Create Excel writer
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        # Write DataFrame to Excel
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        _apply_excel_styling(worksheet, config)

def _apply_excel_styling(worksheet: openpyxl.worksheet.worksheet.Worksheet, config: ExcelLayoutConfig) -> None:
    """Apply styling to an Excel worksheet based on the configuration.
    
    Args:
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to style
        config (ExcelLayoutConfig): Configuration for styling
    """
    # Apply header styling
    if config.bold_headers:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            if config.header_color:
                cell.fill = PatternFill(
                    start_color=config.header_color,
                    end_color=config.header_color,
                    fill_type='solid'
                )
    
    # Set number format
    number_format = config.number_format
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = number_format
    
    # Set row height if specified
    if config.row_height:
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = config.row_height
    
    # Set column widths
    if config.auto_column_width:
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # Apply borders
    if config.horizontal_lines or config.vertical_lines:
        for row in worksheet.iter_rows():
            for cell in row:
                border = Border()
                if config.horizontal_lines:
                    border.top = Side(style='thin')
                    border.bottom = Side(style='thin')
                if config.vertical_lines:
                    border.left = Side(style='thin')
                    border.right = Side(style='thin')
                cell.border = border
    
    # Apply alternating colors if enabled
    if config.alternating_colors:
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(
                        start_color='F0F0F0',
                        end_color='F0F0F0',
                        fill_type='solid'
                    ) 