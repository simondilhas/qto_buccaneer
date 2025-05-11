from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Optional, Union
import pandas as pd
import json, yaml
import zipfile
import os
import ifcopenshell
import io
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from qto_buccaneer._utils.report.excel_styling import ExcelLayoutConfig
import tempfile

class CustomJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder that can handle entity_instance objects and special characters."""
    def default(self, obj):
        if hasattr(obj, '__dict__'):
            return obj.__dict__
        return str(obj)

    def encode(self, obj):
        return super().encode(obj)

@dataclass
class BaseResultBundle:
    """Base class for all BaseResultBundle types with common functionality."""
    dataframe: Optional[pd.DataFrame]
    json: Optional[Dict[str, Any]] = None
    summary: Optional[str] = None
    folderpath: Optional[Path] = None
    ifc_model: Optional['ifcopenshell.file'] = None

    def to_df(self) -> pd.DataFrame:
        """Convert the result bundle to a pandas DataFrame.
        Must be implemented by subclasses."""
        raise NotImplementedError("Subclasses must implement to_df()")

    def to_dict(self) -> Dict[str, Any]:
        """Convert the result bundle to a dictionary format.
        
        Returns:
            Dict[str, Any]: Dictionary containing the data in a format suitable for processing
        """
        if self.json is not None:
            # If we have metadata in the json, use that
            if "metadata" in self.json:
                return self.json["metadata"]
            return self.json
        elif self.dataframe is not None:
            # Convert DataFrame to dictionary format
            elements = {}
            for idx, row in self.dataframe.iterrows():
                element_id = str(idx)
                elements[element_id] = row.to_dict()
            return {"elements": elements}
        else:
            return {}

    def save_json(self, path: Union[str, Path]) -> Path:
        """Save the result bundle to a JSON file."""
        path = Path(path)
        if not path.is_absolute() and self.folderpath is not None:
            path = self.folderpath / path
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(self.json, indent=2, cls=CustomJSONEncoder, ensure_ascii=False), encoding="utf-8")
        return path

    def save_excel(self, path: Union[str, Path], layout_config: Optional[ExcelLayoutConfig] = None) -> Path:
        """Save the result bundle to an Excel file with optional styling.
        
        Args:
            path: String or Path object specifying where to save the Excel file.
                 If relative, will be saved relative to folderpath if set.
            layout_config: Optional ExcelLayoutConfig for styling the Excel file.
                          If None, uses default ExcelLayoutConfig settings.
            
        Returns:
            Path: The path where the Excel file was saved
        """
        if self.dataframe is None:
            self.dataframe = self.to_df()
        path = Path(path)
        if not path.is_absolute() and self.folderpath is not None:
            path = self.folderpath / path
        path.parent.mkdir(parents=True, exist_ok=True)
        
        # Use default config if none provided
        config = layout_config or ExcelLayoutConfig()
        
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            self.dataframe.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            # Freeze panes (first row and first two columns)
            worksheet.freeze_panes = 'C2'
            
            # Add autofilter to header row
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Apply styling based on config
            if config.bold_headers:
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)
                    if config.header_color:
                        cell.fill = PatternFill(
                            start_color=config.header_color,
                            end_color=config.header_color,
                            fill_type='solid'
                        )
            
            # Set number format
            number_format = config.number_format
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = number_format
            
            # Set row height if specified
            if config.row_height:
                for row in worksheet.iter_rows():
                    worksheet.row_dimensions[row[0].row].height = config.row_height
            
            # Set column widths
            if config.auto_column_width:
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            # Apply borders
            if config.horizontal_lines or config.vertical_lines:
                for row in worksheet.iter_rows():
                    for cell in row:
                        border = Border()
                        if config.horizontal_lines:
                            border.top = Side(style='thin')
                            border.bottom = Side(style='thin')
                        if config.vertical_lines:
                            border.left = Side(style='thin')
                            border.right = Side(style='thin')
                        cell.border = border
            
            # Apply alternating colors if enabled
            if config.alternating_colors:
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    if row_idx % 2 == 0:
                        for cell in row:
                            cell.fill = PatternFill(
                                start_color='F0F0F0',
                                end_color='F0F0F0',
                                fill_type='solid'
                            )
        
        return path

@dataclass
class MetadataResultBundle(BaseResultBundle):
    """BaseResultBundle specifically for IFC metadata."""
    def to_df(self) -> pd.DataFrame:
        """Convert metadata JSON to DataFrame."""
        if self.dataframe is None and self.json is not None:
            if "elements" in self.json:
                self.dataframe = pd.DataFrame.from_dict(self.json["elements"], orient='index')
            else:
                self.dataframe = pd.DataFrame([self.json])
        return self.dataframe

@dataclass
class GeometryResultBundle(BaseResultBundle):
    """BaseResultBundle specifically for geometry data."""
    def to_df(self) -> pd.DataFrame:
        """Convert geometry data to DataFrame."""
        if self.dataframe is None and self.json is not None:
            if isinstance(self.json, list):
                self.dataframe = pd.DataFrame(self.json)
            else:
                self.dataframe = pd.DataFrame([self.json])
        return self.dataframe

    def save_geometry(self, path: Union[str, Path]) -> Path:
        """Extract geometry zip file to a directory."""
        if self.json is None or "zip_content" not in self.json:
            raise ValueError("No geometry zip content available")

        path = Path(path)
        if not path.is_absolute() and self.folderpath is not None:
            path = self.folderpath / path
        path.mkdir(parents=True, exist_ok=True)
        
        zip_buffer = io.BytesIO(self.json["zip_content"])
        with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
            zip_file.extractall(path)
        
        return path

    def get_geometry_by_ifc_entity(self, ifc_entity: str) -> Dict[str, Any]:
        """Get geometry data for a specific IFC entity type.
        
        Args:
            ifc_entity: The IFC entity type (e.g., 'IfcSpace', 'IfcWall')
            
        Returns:
            Dict[str, Any]: Dictionary containing geometry data for the specified entity type,
                           with element IDs as keys
            
        Raises:
            ValueError: If no geometry data is available or if the entity type is not found
        """
        if self.json is None or "zip_content" not in self.json:
            raise ValueError("No geometry zip content available")

        # Create a BytesIO object from the zip content
        zip_buffer = io.BytesIO(self.json["zip_content"])
        
        # Read the specific JSON file from the zip
        with zipfile.ZipFile(zip_buffer, 'r') as zip_file:
            try:
                with zip_file.open(f"{ifc_entity}.json") as f:
                    geometry_list = json.loads(f.read().decode('utf-8'))
                    # Convert list to dictionary using element IDs as keys
                    return {item.get('id'): item for item in geometry_list}
            except KeyError:
                raise ValueError(f"No geometry data found for entity type: {ifc_entity}")

@dataclass
class IFCResultBundle(BaseResultBundle):
    """BaseResultBundle specifically for IFC model operations."""
    def save_ifc(self, path: Union[str, Path]) -> Path:
        """Save the IFC model to a file if available.
        
        Args:
            path: String or Path object specifying where to save the IFC file.
                 If relative, will be saved relative to folderpath if set.
            
        Returns:
            Path: The path where the IFC file was saved
            
        Raises:
            ValueError: If no IFC model is available in the result bundle
            
        Note:
            Creates parent directories if they don't exist
        """
        if self.ifc_model is None:
            raise ValueError("Cannot save IFC file: No IFC model available in the result bundle")

        path = Path(path)
        if not path.is_absolute() and self.folderpath is not None:
            path = self.folderpath / path
        path.parent.mkdir(parents=True, exist_ok=True)
        
        self.ifc_model.write(str(path))
        return path

    def get_ifc(self) -> Optional['ifcopenshell.file']:
        """Get the IFC model if available.
        
        Returns:
            Optional[ifcopenshell.file]: The IFC model if available, None otherwise
        """
        return self.ifc_model

@dataclass
class MetricsResultBundle(BaseResultBundle):
    """BaseResultBundle specifically for metrics calculations."""
    def to_df(self) -> pd.DataFrame:
        """Convert metrics data to DataFrame."""
        if self.dataframe is None and self.json is not None:
            if isinstance(self.json, list):
                self.dataframe = pd.DataFrame(self.json)
            else:
                self.dataframe = pd.DataFrame([self.json])
        return self.dataframe
    
    def add_metric(self, metric_name: str, metric_value: float, metric_unit: str, description: str = "") -> None:
        """Add a metric to the metrics DataFrame."""
        if self.dataframe is None:
            self.dataframe = pd.DataFrame()
        new_row = pd.DataFrame([{
            "metric_name": metric_name,
            "value": metric_value,
            "unit": metric_unit,
            "description": description
        }])
        self.dataframe = pd.concat([self.dataframe, new_row], ignore_index=True)

