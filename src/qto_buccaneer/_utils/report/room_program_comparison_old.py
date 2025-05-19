from typing import Optional
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from qto_buccaneer.report import ExcelLayoutConfig
import os




def export_room_program_comparison_old(
    df: pd.DataFrame,
    output_path: str,
    layout_config: Optional[ExcelLayoutConfig] = None
) -> pd.DataFrame:
    """
    Helper function to export room program comparison to Excel with formatting.
    Used by room_program_comparison() to handle Excel export.
    
    Args:
        df (pd.DataFrame): Room program comparison DataFrame
        output_path (str): Full path where the Excel file should be saved
        layout_config (Optional[ExcelLayoutConfig]): Configuration for Excel layout.
            If None, default settings will be used.
            
    Returns:
        pd.DataFrame: The exported DataFrame
        
    Note:
        This is an internal helper function and should not be called directly.
        Use room_program_comparison() instead.
    """
    if df.empty:
        print("Warning: Input DataFrame is empty!")
        return df
        
    # Make sure the output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir:  # Only create directory if path contains a directory component
        os.makedirs(output_dir, exist_ok=True)
        
    try:
        # Use provided config or create default one
        config = layout_config or ExcelLayoutConfig(
            horizontal_lines=True,
            vertical_lines=True,
            bold_headers=True,
            auto_column_width=True,
            alternating_colors=True,
            number_format='#,##0.00'
        )
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Room Program Comparison')
            worksheet = writer.sheets['Room Program Comparison']
            
            # Auto-adjust column widths
            if config.auto_column_width:
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    adjusted_width = max_length + 2
                    column_letter = get_column_letter(idx + 1)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height if specified
            if config.row_height:
                for row in range(1, len(df) + 2):
                    worksheet.row_dimensions[row].height = config.row_height
            
            # Format cells
            for row in range(1, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Add borders based on config
                    borders = {}
                    if config.horizontal_lines:
                        borders['bottom'] = openpyxl.styles.Side(style='thin')
                    if config.vertical_lines:
                        borders['right'] = openpyxl.styles.Side(style='thin')
                    if borders:
                        cell.border = openpyxl.styles.Border(**borders)
                    
                    # Format numbers (skip first column - Room Type)
                    if row > 1 and col > 1:
                        try:
                            float(cell.value)  # Check if value is numeric
                            # Use percentage format for % columns
                            if df.columns[col-1].startswith('%'):
                                cell.number_format = '0.00%'
                            else:
                                cell.number_format = config.number_format
                        except (TypeError, ValueError):
                            pass
                    
                    # Format headers
                    if row == 1 and config.bold_headers:
                        cell.font = openpyxl.styles.Font(bold=True)
                        if config.header_color:
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color=config.header_color,
                                end_color=config.header_color,
                                fill_type='solid'
                            )
                    
                    # Format totals row (last row)
                    if row == len(df) + 1:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color=config.header_color,
                            end_color=config.header_color,
                            fill_type='solid'
                        )
                    
                    # Apply alternating colors
                    elif config.alternating_colors and row > 1 and row <= len(df):
                        if row % 2 == 0:  # Even rows
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color='F5F5F5',  # Light gray
                                end_color='F5F5F5',
                                fill_type='solid'
                            )

        print(f"Excel file successfully created at: {output_path}")
    except Exception as e:
        print(f"Error creating Excel file: {str(e)}")
        raise
        
    return df
