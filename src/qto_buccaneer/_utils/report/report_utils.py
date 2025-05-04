from qto_buccaneer.report import ExcelLayoutConfig
from typing import Optional
import pandas as pd
import openpyxl
import os


def _export_to_excel(
    df: pd.DataFrame, 
    output_dir: str,
    filename_suffix: str = "metrics",
    building_name: Optional[str] = None,
    layout_config: Optional[ExcelLayoutConfig] = None
) -> str:
    """Helper function to export a DataFrame to a new Excel file with optional styling.
    Used by other functions to handle Excel export.
    
    Args:
        df (pd.DataFrame): DataFrame to export
        output_dir (str): Directory where the Excel file should be saved
        filename_suffix (str): Suffix to add to the filename (default: "metrics")
        building_name (Optional[str]): Name of the building to include in filename
        layout_config (Optional[ExcelLayoutConfig]): Configuration for Excel styling. 
            If None, uses default ExcelLayoutConfig settings.
            
    Returns:
        str: The path to the exported Excel file
        
    Note:
        This is an internal helper function and should not be called directly.
        Use the appropriate public function instead.
    """
    if df.empty:
        return df
        
    # Use default config if none provided
    if layout_config is None:
        layout_config = ExcelLayoutConfig()
    
    # Create output directory if it doesn't exist
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    
    # Construct filename
    filename_parts = []
    if building_name:
        filename_parts.append(building_name)
    filename_parts.append(filename_suffix)
    filename = "_".join(filename_parts) + ".xlsx"
    
    # Create full output path
    output_path = os.path.join(output_dir, filename)
    
    # Create Excel writer with the full path
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write DataFrame to Excel
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Apply styling based on config
        if layout_config.bold_headers:
            for cell in worksheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                if layout_config.header_color:
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color=layout_config.header_color,
                        end_color=layout_config.header_color,
                        fill_type='solid'
                    )
        
        # Set number format
        number_format = layout_config.number_format
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format
        
        # Set row height if specified
        if layout_config.row_height:
            for row in worksheet.iter_rows():
                worksheet.row_dimensions[row[0].row].height = layout_config.row_height
        
        # Set column widths
        if layout_config.auto_column_width:
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Apply borders
        if layout_config.horizontal_lines or layout_config.vertical_lines:
            for row in worksheet.iter_rows():
                for cell in row:
                    border = openpyxl.styles.Border()
                    if layout_config.horizontal_lines:
                        border.top = openpyxl.styles.Side(style='thin')
                        border.bottom = openpyxl.styles.Side(style='thin')
                    if layout_config.vertical_lines:
                        border.left = openpyxl.styles.Side(style='thin')
                        border.right = openpyxl.styles.Side(style='thin')
                    cell.border = border
        
        # Apply alternating colors if enabled
        if layout_config.alternating_colors:
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color='F0F0F0',
                            end_color='F0F0F0',
                            fill_type='solid'
                        )

    return output_path