import pandas as pd
import os
from pathlib import Path
from typing import Optional, Dict, Any
from jinja2 import Environment, FileSystemLoader
import subprocess
from dataclasses import dataclass
import openpyxl
from openpyxl.utils import get_column_letter

@dataclass
class ExcelLayoutConfig:
    """Configuration for Excel export layout."""
    horizontal_lines: bool = True
    vertical_lines: bool = False
    bold_headers: bool = True
    auto_column_width: bool = True
    row_height: Optional[float] = None
    alternating_colors: bool = False
    number_format: str = '#,##0.00'
    header_color: str = 'E0E0E0'  # Light gray
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert config to dictionary."""
        return {k: v for k, v in self.__dict__.items()}

def export_to_excel(
    df: pd.DataFrame, 
    path: str, 
    config: Optional[ExcelLayoutConfig] = None
) -> None:
    """Export a DataFrame to a new Excel file with optional styling.
    
    Args:
        df (pd.DataFrame): DataFrame to export
        path (str): Path where the Excel file should be saved
        config (Optional[ExcelLayoutConfig]): Configuration for Excel styling. 
            If None, uses default ExcelLayoutConfig settings.
    """
    if df.empty:
        return
        
    # Use default config if none provided
    if config is None:
        config = ExcelLayoutConfig()
    
    # Create Excel writer
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        # Write DataFrame to Excel
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Apply styling based on config
        if config.bold_headers:
            for cell in worksheet[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                if config.header_color:
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color=config.header_color,
                        end_color=config.header_color,
                        fill_type='solid'
                    )
        
        # Set number format
        number_format = config.number_format
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format
        
        # Set row height if specified
        if config.row_height:
            for row in worksheet.iter_rows():
                worksheet.row_dimensions[row[0].row].height = config.row_height
        
        # Set column widths
        if config.auto_column_width:
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Apply borders
        if config.horizontal_lines or config.vertical_lines:
            for row in worksheet.iter_rows():
                for cell in row:
                    border = openpyxl.styles.Border()
                    if config.horizontal_lines:
                        border.top = openpyxl.styles.Side(style='thin')
                        border.bottom = openpyxl.styles.Side(style='thin')
                    if config.vertical_lines:
                        border.left = openpyxl.styles.Side(style='thin')
                        border.right = openpyxl.styles.Side(style='thin')
                    cell.border = border
        
        # Apply alternating colors if enabled
        if config.alternating_colors:
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color='F0F0F0',
                            end_color='F0F0F0',
                            fill_type='solid'
                        )

def generate_pdf_report(
    project_data: dict, 
    template_path: str = "template.tex", 
    output_path: str = "output.tex") -> None:
    
    """Generate a PDF report from a LaTeX template using project data.
    
    Args:
        project_data (dict): Dictionary containing data to be rendered in the template
        template_path (str, optional): Path to the LaTeX template file. Defaults to "template.tex"
        output_path (str, optional): Path for the output tex file. Defaults to "output.tex"
    """
    # Get the directory containing the template
    template_dir = str(Path(template_path).parent)
    template_name = Path(template_path).name
    
    # Set up Jinja environment with the correct directory
    env = Environment(loader=FileSystemLoader(template_dir))
    template = env.get_template(template_name)
    rendered = template.render(**project_data)

    # Write rendered template to file
    with open(output_path, "w") as f:
        f.write(rendered)

    # Generate PDF using pdflatex
    subprocess.run(["pdflatex", output_path])

def create_project_comparison_df(df: pd.DataFrame, metrics: Optional[list[str]] = None) -> pd.DataFrame:
    """
    Create a comparison DataFrame with projects as rows and metrics as columns.

    Args:
        df (pd.DataFrame): Input DataFrame containing metrics data
        metrics (Optional[list[str]]): List of metric names to include in the comparison.
            If None, all metrics will be included.

    Returns:
        pd.DataFrame: DataFrame with projects as rows and metrics as columns
    """
    required_columns = {'file_name', 'metric_name', 'unit', 'value'}
    missing_columns = required_columns - set(df.columns)
    
    if missing_columns:
        print(f"Warning: Missing required columns: {missing_columns}")  # Debug print
        print(f"Available columns: {df.columns.tolist()}")  # Debug print
        return pd.DataFrame()
        
    try:
        # Filter metrics if a list is provided
        if metrics is not None:
            df = df[df['metric_name'].isin(metrics)].copy()
            
            if df.empty:
                return pd.DataFrame()
        
        # Create a copy to avoid modifying the original DataFrame
        df = df.copy()
        
        # Clean up project names by removing the suffix
        df['file_name'] = df['file_name'].str.replace('_abstractBIM_sp_enriched.ifc', '')
        
        # Create metric names with units
        df['metric_with_unit'] = df['metric_name'] + ' [' + df['unit'] + ']'
        
        pivot_df = df.pivot(
            index='file_name',
            columns='metric_with_unit',
            values='value'
        )
        
        # Only sort if no specific metrics order was provided
        if metrics is None:
            pivot_df = pivot_df.sort_index(axis=1)
        else:
            # Reorder columns based on the provided metrics order
            metric_order = [m + ' [' + df[df['metric_name'] == m]['unit'].iloc[0] + ']' for m in metrics]
            pivot_df = pivot_df[metric_order]
            
        pivot_df = pivot_df.reset_index()
        pivot_df = pivot_df.rename(columns={'file_name': 'Project'})
        
        return pivot_df
        
    except Exception as e:
        return pd.DataFrame()



def export_project_comparison_excel(
    df: pd.DataFrame, 
    output_path: str, 
    metrics: Optional[list[str]] = None,
    layout_config: Optional[ExcelLayoutConfig] = None
) -> pd.DataFrame:
    """
    Create and export project comparison to Excel file with customizable formatting.

    Args:
        df (pd.DataFrame): Input DataFrame containing metrics data
        output_path (str): Path where the Excel file should be saved
        metrics (Optional[list[str]]): List of metric names to include in the comparison
        layout_config (Optional[ExcelLayoutConfig]): Configuration for Excel layout.
            If None, default settings will be used.
            
    Returns:
        pd.DataFrame: The comparison DataFrame that was exported
    """
    comparison_df = create_project_comparison_df(df, metrics)
    
    print("Input DataFrame shape:", df.shape)  # Debug print
    print("Comparison DataFrame shape:", comparison_df.shape)  # Debug print
    print("Metrics:", metrics)  # Debug print
    
    if comparison_df.empty:
        print("Warning: Comparison DataFrame is empty!")  # Debug print
        return comparison_df
        
    # Make sure the output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    try:
        config = layout_config or ExcelLayoutConfig()
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            comparison_df.to_excel(writer, index=False, sheet_name='Comparison')
            worksheet = writer.sheets['Comparison']
            
            # Auto-adjust column widths if enabled
            if config.auto_column_width:
                for idx, col in enumerate(comparison_df.columns):
                    max_length = max(
                        comparison_df[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    adjusted_width = max_length + 2
                    column_letter = get_column_letter(idx + 1)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height if specified
            if config.row_height:
                for row in range(1, len(comparison_df) + 2):
                    worksheet.row_dimensions[row].height = config.row_height
            
            # Add gridlines
            for row in range(1, len(comparison_df) + 2):
                for col in range(1, len(comparison_df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    borders = {}
                    
                    if config.horizontal_lines:
                        borders['bottom'] = openpyxl.styles.Side(style='thin')
                    if config.vertical_lines:
                        borders['right'] = openpyxl.styles.Side(style='thin')
                    
                    if borders:
                        cell.border = openpyxl.styles.Border(**borders)
                    
                    # Apply number format to numeric cells
                    if row > 1 and col > 1:  # Skip header row and project column
                        try:
                            float(cell.value)  # Check if value is numeric
                            cell.number_format = config.number_format
                        except (TypeError, ValueError):
                            pass
            
            # Format headers
            if config.bold_headers:
                for col in range(1, len(comparison_df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = openpyxl.styles.Font(bold=True)
                    
                    if config.header_color:
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color=config.header_color,
                            end_color=config.header_color,
                            fill_type='solid'
                        )
            
            # Apply alternating colors if enabled
            if config.alternating_colors:
                for row in range(2, len(comparison_df) + 2, 2):  # Start after header
                    for col in range(1, len(comparison_df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color='F5F5F5',  # Light gray
                            end_color='F5F5F5',
                            fill_type='solid'
                        )

        print(f"Excel file successfully created at: {output_path}")  # Debug print
    except Exception as e:
        print(f"Error creating Excel file: {str(e)}")  # Debug print
        raise
        
    return comparison_df

def room_program_comparison(
    target_excel_path: str,
    ifc_loader,
    room_name_column: str = "LongName",
    target_count_column: str = "Target Count",
    target_area_column: str = "Target Area/Room",
    output_path: Optional[str] = None,
    layout_config: Optional[ExcelLayoutConfig] = None
) -> pd.DataFrame:
    """
    Create a comparison between target room program and actual IFC spaces.
    
    Args:
        target_excel_path: Path to Excel file containing target room program
        ifc_loader: Instance of IfcLoader with loaded IFC model
        room_name_column: Column name in target Excel for room names (default: "Room Type")
        target_count_column: Column name for target room count (default: "Target Count")
        target_area_column: Column name for target room area (default: "Target Area")
        output_path: Optional path to save Excel report
        layout_config: Optional ExcelLayoutConfig for custom formatting
        
    Returns:
        pd.DataFrame: Comparison table with target vs actual metrics
    """
    # Load target room program
    try:
        target_df = pd.read_excel(target_excel_path)
        print(f"Loaded target Excel file. Columns found: {target_df.columns.tolist()}")  # Debug print
        
        # Verify required columns exist
        missing_columns = []
        for col, col_name in [
            (room_name_column, "room name"),
            (target_count_column, "target count"),
            (target_area_column, "target area")
        ]:
            if col not in target_df.columns:
                missing_columns.append(f"{col_name} ({col})")
        
        if missing_columns:
            raise ValueError(f"Missing required columns in target Excel: {', '.join(missing_columns)}")
            
    except Exception as e:
        print(f"Error loading target Excel file: {str(e)}")
        return pd.DataFrame()
        
    # Get actual spaces from IFC
    try:
        spaces_df = ifc_loader.get_space_information()
        print(f"Loaded spaces from IFC. Found {len(spaces_df)} spaces.")  # Debug print
        
        if spaces_df.empty:
            raise ValueError("No spaces found in IFC model")
            
        # Verify required IFC data columns
        if 'LongName' not in spaces_df.columns:
            raise ValueError("IFC spaces missing 'LongName' attribute")
        if 'Qto_SpaceBaseQuantities.NetFloorArea' not in spaces_df.columns:
            raise ValueError("IFC spaces missing 'NetFloorArea' quantity")
            
    except Exception as e:
        print(f"Error processing IFC spaces: {str(e)}")
        return pd.DataFrame()
    
    # Initialize result DataFrame
    result = pd.DataFrame()
    
    try:
        # Process each room type from target program
        data = []
        for _, row in target_df.iterrows():
            room_name = row[room_name_column]
            target_count = float(row[target_count_column])  # Convert to float for safety
            target_area = float(row[target_area_column])
            
            print(f"Processing room type: {room_name}")  # Debug print
            
            # Get actual spaces matching this room name
            actual_spaces = spaces_df[spaces_df['LongName'] == room_name]
            actual_count = len(actual_spaces)
            
            # Sum up actual areas
            actual_total_area = actual_spaces['Qto_SpaceBaseQuantities.NetFloorArea'].sum()
            
            print(f"Found {actual_count} spaces with total area {actual_total_area}")  # Debug print
            
            # Calculate metrics
            target_total_area = target_count * target_area
            avg_area_per_room = actual_total_area / actual_count if actual_count > 0 else 0
            
            count_diff = actual_count - target_count
            count_diff_pct = (count_diff / target_count * 100) if target_count > 0 else 0
            
            area_diff = actual_total_area - target_total_area
            area_diff_pct = (area_diff / target_total_area * 100) if target_total_area > 0 else 0
            
            data.append({
                'Room Type': room_name,
                'Target Count': target_count,
                'Target sqm/room': target_area,
                'Target Total sqm': target_total_area,
                'Actual Count': actual_count,
                'Actual Total sqm': actual_total_area,
                'Avg sqm/room': avg_area_per_room,
                'Count Diff': count_diff,
                '% Count Diff': count_diff_pct,
                'Area Diff': area_diff,
                '% Area Diff': area_diff_pct
            })
            
        result = pd.DataFrame(data)
        
        if result.empty:
            print("No data was processed - empty result DataFrame")
            return pd.DataFrame()
            
        # Add totals row
        totals = {
            'Room Type': 'TOTAL',
            'Target Count': result['Target Count'].sum(),
            'Target Total sqm': result['Target Total sqm'].sum(),
            'Actual Count': result['Actual Count'].sum(),
            'Actual Total sqm': result['Actual Total sqm'].sum(),
            'Count Diff': result['Count Diff'].sum(),
            'Area Diff': result['Area Diff'].sum()
        }
        
        # Calculate weighted averages for percentages
        total_target_count = result['Target Count'].sum()
        total_target_area = result['Target Total sqm'].sum()
        
        if total_target_count > 0:
            totals['% Count Diff'] = (totals['Count Diff'] / total_target_count * 100)
        if total_target_area > 0:
            totals['% Area Diff'] = (totals['Area Diff'] / total_target_area * 100)
            
        # Calculate overall average sqm/room
        if totals['Actual Count'] > 0:
            totals['Avg sqm/room'] = totals['Actual Total sqm'] / totals['Actual Count']
        else:
            totals['Avg sqm/room'] = 0
            
        totals['Target sqm/room'] = totals['Target Total sqm'] / totals['Target Count'] if totals['Target Count'] > 0 else 0
        
        # Append totals row
        result = pd.concat([result, pd.DataFrame([totals])], ignore_index=True)
        
        # Export to Excel if output path is provided
        if output_path:
            result = export_room_program_comparison(
                df=result,
                output_path=output_path,
                layout_config=layout_config
            )
        
        return result
        
    except Exception as e:
        print(f"Error creating comparison: {str(e)}")
        import traceback
        traceback.print_exc()  # Print full stack trace
        return pd.DataFrame()

def export_room_program_comparison(
    df: pd.DataFrame,
    output_path: str,
    layout_config: Optional[ExcelLayoutConfig] = None
) -> pd.DataFrame:
    """
    Export room program comparison to Excel with formatting.
    
    Args:
        df (pd.DataFrame): Room program comparison DataFrame
        output_path (str): Path where the Excel file should be saved
        layout_config (Optional[ExcelLayoutConfig]): Configuration for Excel layout.
            If None, default settings will be used.
            
    Returns:
        pd.DataFrame: The exported DataFrame
    """
    if df.empty:
        print("Warning: Input DataFrame is empty!")
        return df
        
    # Make sure the output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    try:
        # Use provided config or create default one
        config = layout_config or ExcelLayoutConfig(
            horizontal_lines=True,
            vertical_lines=True,
            bold_headers=True,
            auto_column_width=True,
            alternating_colors=True,
            number_format='#,##0.00'
        )
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Room Program Comparison')
            worksheet = writer.sheets['Room Program Comparison']
            
            # Auto-adjust column widths
            if config.auto_column_width:
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(str(col))
                    )
                    adjusted_width = max_length + 2
                    column_letter = get_column_letter(idx + 1)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Set row height if specified
            if config.row_height:
                for row in range(1, len(df) + 2):
                    worksheet.row_dimensions[row].height = config.row_height
            
            # Format cells
            for row in range(1, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Add borders based on config
                    borders = {}
                    if config.horizontal_lines:
                        borders['bottom'] = openpyxl.styles.Side(style='thin')
                    if config.vertical_lines:
                        borders['right'] = openpyxl.styles.Side(style='thin')
                    if borders:
                        cell.border = openpyxl.styles.Border(**borders)
                    
                    # Format numbers (skip first column - Room Type)
                    if row > 1 and col > 1:
                        try:
                            float(cell.value)  # Check if value is numeric
                            # Use percentage format for % columns
                            if df.columns[col-1].startswith('%'):
                                cell.number_format = '0.00%'
                            else:
                                cell.number_format = config.number_format
                        except (TypeError, ValueError):
                            pass
                    
                    # Format headers
                    if row == 1 and config.bold_headers:
                        cell.font = openpyxl.styles.Font(bold=True)
                        if config.header_color:
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color=config.header_color,
                                end_color=config.header_color,
                                fill_type='solid'
                            )
                    
                    # Format totals row (last row)
                    if row == len(df) + 1:
                        cell.font = openpyxl.styles.Font(bold=True)
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color=config.header_color,
                            end_color=config.header_color,
                            fill_type='solid'
                        )
                    
                    # Apply alternating colors
                    elif config.alternating_colors and row > 1 and row <= len(df):
                        if row % 2 == 0:  # Even rows
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color='F5F5F5',  # Light gray
                                end_color='F5F5F5',
                                fill_type='solid'
                            )

        print(f"Excel file successfully created at: {output_path}")
    except Exception as e:
        print(f"Error creating Excel file: {str(e)}")
        raise
        
    return df

# Example usage:
"""
# Use default settings
export_project_comparison_excel(df, 'comparison.xlsx')

# Custom layout
config = ExcelLayoutConfig(
    horizontal_lines=True,
    vertical_lines=True,
    bold_headers=True,
    auto_column_width=True,
    row_height=20,
    alternating_colors=True,
    number_format='#,##0.00',
    header_color='E0E0E0'
)
export_project_comparison_excel(df, 'comparison.xlsx', layout_config=config)

# Minimal layout
minimal_config = ExcelLayoutConfig(
    horizontal_lines=False,
    vertical_lines=False,
    bold_headers=True,
    auto_column_width=True
)
export_project_comparison_excel(df, 'comparison.xlsx', layout_config=minimal_config)
""" 