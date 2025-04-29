from qto_buccaneer.reports import ExcelLayoutConfig
from typing import Optional
import pandas as pd
import openpyxl
import os
import yaml
import traceback
from dataclasses import dataclass
from pathlib import Path
from openpyxl.utils import get_column_letter

class RoomComparisonResult:
    """Class to hold and format room comparison results."""
    
    def __init__(self, detailed_df: pd.DataFrame, ifc_rooms: set, excel_rooms: set):
        self.detailed_df = detailed_df
        self.ifc_rooms = ifc_rooms
        self.excel_rooms = excel_rooms
        
        # Calculate summary statistics
        self.total_target_rooms = len(excel_rooms)
        self.total_ifc_rooms = len(ifc_rooms)
        self.matching_rooms = len(ifc_rooms.intersection(excel_rooms))
        self.missing_rooms = list(excel_rooms - ifc_rooms)
        self.extra_rooms = list(ifc_rooms - excel_rooms)
        
        # Determine status
        if not self.missing_rooms and not self.extra_rooms:
            self.status = "passed"
        elif self.missing_rooms:
            self.status = "failed-rooms missing in ifc"
        else:
            self.status = "failed-rooms added in ifc"
    
    def to_yaml(self) -> dict:
        """Generate YAML summary of the comparison."""
        summary = {
            "type": "room_comparison",
            "status": self.status,
            "summary": f"{self.matching_rooms} of {self.total_target_rooms} rooms found in IFC",
            "target": {
                "total_rooms": self.total_target_rooms,
                "found_rooms": self.matching_rooms,
                "missing_rooms": len(self.missing_rooms)
            },
            "ifc": {
                "total_rooms": self.total_ifc_rooms,
                "matching_rooms": self.matching_rooms,
                "extra_rooms": len(self.extra_rooms)
            }
        }
        
        # Add issue details if there are any
        if self.status != "passed":
            summary["issues"] = {}
            if self.missing_rooms:
                summary["issues"]["missing_rooms"] = [
                    {"name": room} for room in sorted(self.missing_rooms)
                ]
            if self.extra_rooms:
                # Get GlobalIds for extra rooms from the detailed DataFrame
                extra_rooms_data = []
                for room in sorted(self.extra_rooms):
                    room_data = self.detailed_df[
                        (self.detailed_df["Room Name"].str.lower() == room) & 
                        (self.detailed_df["Status"] == "Only in IFC")
                    ]
                    if not room_data.empty:
                        extra_rooms_data.append({
                            "global_id": room_data["GlobalId"].iloc[0],
                            "LongName": room,
                        })
                    else:
                        extra_rooms_data.append({"name": room})
                summary["issues"]["extra_rooms"] = extra_rooms_data
        
        return summary
    
    def to_dict(self) -> dict:
        """Get all comparison data as a dictionary."""
        return {
            "detailed_df": self.detailed_df.to_dict(),
            "summary": self.to_yaml(),
            "status": self.status,
            "statistics": {
                "total_target_rooms": self.total_target_rooms,
                "total_ifc_rooms": self.total_ifc_rooms,
                "matching_rooms": self.matching_rooms,
                "missing_rooms_count": len(self.missing_rooms),
                "extra_rooms_count": len(self.extra_rooms)
            }
        }
    
    def to_excel(self, output_path: str, layout_config: Optional[ExcelLayoutConfig] = None) -> None:
        """Export detailed comparison to Excel with formatting."""
        if self.detailed_df.empty:
            print("Warning: No data to export to Excel!")
            return
            
        # Make sure the output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
            
        try:
            # Use provided config or create default one
            config = layout_config or ExcelLayoutConfig(
                horizontal_lines=True,
                vertical_lines=True,
                bold_headers=True,
                auto_column_width=True
            )
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                self.detailed_df.to_excel(writer, index=False, sheet_name='Room Name Comparison')
                worksheet = writer.sheets['Room Name Comparison']
                
                # Apply styling based on config
                if config.bold_headers:
                    for cell in worksheet[1]:
                        cell.font = openpyxl.styles.Font(bold=True)
                        if config.header_color:
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color=config.header_color,
                                end_color=config.header_color,
                                fill_type='solid'
                            )
                
                # Auto-adjust column widths
                if config.auto_column_width:
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
                        
        except Exception as e:
            print(f"Error exporting to Excel: {str(e)}")
            import traceback
            traceback.print_exc()



class SafeLoader(yaml.SafeLoader):
    """Custom YAML loader that handles both scalar and sequence nodes."""
    def construct_scalar(self, node):
        if isinstance(node, yaml.SequenceNode):
            return [self.construct_scalar(child) for child in node.value]
        return super().construct_scalar(node)


@dataclass
class BuildingComparison:
    """Main data structure for building comparison results."""
    merged_df: pd.DataFrame
    return_values: list[str] 
    area_target_column: str
    area_actual_column: str
    
    def __post_init__(self):
        """Initialize the comparison."""
        pass
    
    def to_summary_yaml(self) -> dict:
        """Generate YAML summary of the comparison."""
        # Calculate statistics from merged DataFrame
        total_items = len(self.merged_df)
        items_within_tolerance = len(self.merged_df[self.merged_df['status'] == 'within_tolerance'])
        missing_items = len(self.merged_df[self.merged_df['status'] == 'missing'])
        extra_items = len(self.merged_df[self.merged_df['status'] == 'extra'])
        
        # Calculate total areas
        total_target_area = float(self.merged_df[self.area_target_column].sum())
        total_actual_area = float(self.merged_df[self.area_actual_column].sum())
        total_area_diff = float(total_actual_area - total_target_area)
        total_area_diff_pct = float((total_area_diff / total_target_area * 100) if total_target_area > 0 else 0)
        
        # Determine status
        if missing_items > 0:
            status = "failed"
        elif extra_items > 0:
            status = "warning"
        elif items_within_tolerance == total_items:
            status = "passed"
        else:
            status = "warning"
        
        # Get problematic items
        #problematic_df = self.merged_df[self.merged_df['status'] != 'within_tolerance']
        #problematic_items = [
        #    {
        #        "category": str(row['target_key']),
        #        "status": str(row['status']),
        #        "name": str(row.get('name_target', row.get('name_actual', ''))),
        #        "target": float(row['target_area']),
        #        "actual": float(row['actual_area']),
        #        "difference_pct": float(row['area_diff_pct'])
        #    }
        #    for _, row in problematic_df.iterrows()
        #]
        
        summary = {
            "name": "Target Program vs Actual Program",
            "status": str(status),
            "overview": {
                "total_items": int(total_items),
                "compliance_rate": float((items_within_tolerance / total_items * 100) if total_items > 0 else 0),
                "total_area": {
                    "target": float(total_target_area),
                    "actual": float(total_actual_area),
                    "difference_pct": float(total_area_diff_pct)
                }
            },
            "issues": {
                "missing": int(missing_items),
                "extra": int(extra_items),
                "out_of_tolerance": int(total_items - items_within_tolerance - missing_items - extra_items)
            },
            #"problematic_items": sorted(
            #    problematic_items,
            #    key=lambda x: abs(x["difference_pct"]),
            #    reverse=True
            #)[:10]
        }
        
        return summary
    
    def to_dict(self) -> dict:
        """Convert the comparison data to a dictionary."""
        return self.to_dataframe().to_dict(orient='records')
    
    def to_dataframe(self,) -> pd.DataFrame:
        """Return the merged DataFrame with the spcified columns."""
        return self.merged_df[self.return_values]

def _load_filter_config(filter_dir: str) -> dict:
    """Load filter configuration from file or string."""
    if Path(filter_dir).exists():
        with open(filter_dir, 'r') as f:
            return yaml.load(f, Loader=SafeLoader)
    return yaml.load(filter_dir, Loader=SafeLoader)

def _get_check_config(filter_config: dict) -> dict:
    """Extract check configuration from filter config."""
    checks = filter_config.get('checks', [])
    if not checks:
        raise ValueError("No checks configuration found in YAML file")
    return checks[0]


def _calculate_differences(
    merged_df: pd.DataFrame,
    tolerance: float,  # e.g. 10.0 for ±10%
    target_area_column: str,
    actual_area_column: str,
    key_target_column: str,
    key_actual_column: str
) -> pd.DataFrame:
    """Calculate area differences and determine status based on percentage tolerance."""

    # Fill missing values
    merged_df[actual_area_column] = merged_df[actual_area_column].fillna(0.0)
    merged_df[target_area_column] = merged_df[target_area_column].fillna(0.0)

    # Compute area difference
    merged_df['area_diff'] = merged_df[actual_area_column] - merged_df[target_area_column]

    # Compute percentage difference based on target area, safely
    merged_df['area_diff_pct'] = 0.0
    mask = merged_df[target_area_column] > 0
    merged_df.loc[mask, 'area_diff_pct'] = (
        (merged_df.loc[mask, 'area_diff'].abs() / merged_df.loc[mask, target_area_column]) * 100
    )

    # Determine status
    def get_status(row, key_target_column, key_actual_column, actual_area_column, target_area_column, tolerance):
        actual_area = row.get(actual_area_column, 0.0)
        target_area = row.get(target_area_column, 0.0)
        key_target_column = row.get(key_target_column, None)

        # Small value threshold
        EPSILON = 0.001  # treat < 0.001 as zero

        # 1. Missing → Target exists, actual missing
        if target_area > EPSILON and actual_area < EPSILON:
            return 'missing'

        # 2. Project-specific → Target exists (LongName present), no area planned, but actual area > 0
        if pd.notna(key_target_column) and target_area < EPSILON and actual_area > EPSILON:
            return 'project_specific'

        # 3. Extra space → No target LongName, but actual area > 0
        if pd.isna(key_target_column) and actual_area > EPSILON:
            return 'extra_space'

        # 4. Within tolerance
        if target_area > EPSILON:
            lower_bound = target_area * (1 - tolerance / 100.0)
            upper_bound = target_area * (1 + tolerance / 100.0)
            if lower_bound <= actual_area <= upper_bound:
                return 'within_tolerance'

        # 5. Otherwise: Out of tolerance
        return 'out_of_tolerance'

    merged_df['status'] = merged_df.apply(
    lambda row: get_status(
        row,
        key_target_column=key_target_column,
        key_actual_column=key_actual_column,
        actual_area_column=actual_area_column,
        target_area_column=target_area_column,
        tolerance=tolerance
    ),
    axis=1
)

    return merged_df



def _save_results(comparison: BuildingComparison, output_path: Path, building_name: str) -> None:
    """Save comparison results to files."""
    # Save YAML summary
    summary = comparison.to_summary_yaml()
    with open(output_path / f"{building_name}_summary.yaml", 'w') as f:
        yaml.dump(summary, f)
    
    # Save DataFrame
    df = comparison.to_dataframe()
    df.to_excel(output_path / f"{building_name}_comparison.xlsx", index=False)